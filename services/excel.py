import threading
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font

from config import EXCEL_PATH, MONTH_NAMES

_lock = threading.Lock()

HEADERS = ["Дата", "Магазин / получатель", "Сумма, RSD", "Категория", "Комментарий"]


def _sheet_name(date_str: str) -> str:
    """'DD.MM' -> 'апрель 2026'"""
    month_num = int(date_str.split(".")[1])
    year = datetime.now().year
    return f"{MONTH_NAMES[month_num]} {year}"


def _get_or_create_sheet(wb, date_str: str):
    name = _sheet_name(date_str)
    if name in wb.sheetnames:
        return wb[name]

    # create new sheet before the summary sheet
    summary_idx = None
    for i, sn in enumerate(wb.sheetnames):
        if "месяц" in sn.lower():
            summary_idx = i
            break

    if summary_idx is not None:
        ws = wb.create_sheet(name, summary_idx)
    else:
        ws = wb.create_sheet(name)

    # write headers matching existing style
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    return ws


def append_transaction(date: str, merchant: str, amount: int, category: str, comment: str = "") -> None:
    with _lock:
        wb = load_workbook(EXCEL_PATH)
        ws = _get_or_create_sheet(wb, date)

        # find first empty row (skip header)
        next_row = ws.max_row + 1
        ws.cell(next_row, 1, date)
        ws.cell(next_row, 2, merchant)
        ws.cell(next_row, 3, amount)
        ws.cell(next_row, 4, category)
        if comment:
            ws.cell(next_row, 5, comment)

        wb.save(EXCEL_PATH)
